# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------
# BaixaFaltes50
#
# Script per descarregar, processar i analitzar les faltes d'assistència
# de l'alumnat, generant informes i resums per al seguiment.
#
# Autor: Josem
# Versió: 50
# Data: 1 de febrer de 2026
# Llicència: Creative Commons BY-NC-SA 4.0
# ----------------------------------------------------------------------
print("⏳ [1/5] Iniciant script...")

import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog
import os
import time
import shutil
import threading
import unicodedata 
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import numpy as np
import traceback
import webbrowser
from tkinter import font
import base64

# Llibreries correu
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

print("⏳ [2/5] Important llibreries gràfiques i de dades... FET.")
print("⏳ [3/5] Important driver de Chrome (això pot trigar una mica)...")

import undetected_chromedriver as uc

print("⏳ [4/5] Driver importat. Definint classes...")

# --- CONFIGURACIÓ GLOBAL ---
URL_BASE_FALTES = "https://insramonturroidarder.ieduca.com/pop/informe_assistencia3.php"
PARAMS_FIXOS = "&hora1=&hora2=&grup=&grups2=&condicio=&major=&professor=&alumne=&id_cicle=16&grup_cicle=&7=1&6=1&15=1&4=1&9=1&8=1"

# NOMS DE FITXERS
NOM_FITXER_MESTRE_XLSX = "dades.xlsx"
NOM_FITXER_MESTRE_ODS = "dades.ods"
NOM_FITXER_EXPEDIENTS = "Expedients2526.ods"
NOM_FULL_EXPEDIENTS = "Curs25-26"

class AplicacioFaltes:
    def __init__(self, root):
        print("   -> Iniciant script...")
        self.root = root
        self.root.title("Gestor Faltes (V50)")
        self.root.geometry("1100x900")
        
        self.driver = None
        self.carpeta_desti = tk.StringVar()
        self.var_versio_chrome = tk.StringVar(value="144") 
        self.carpeta_temp = os.path.join(os.getcwd(), "temp_faltes")
        
        self.var_email_origen = tk.StringVar(value="cap.estudis@iesmalgrat.cat") 
        self.var_password = tk.StringVar(value="d3RvYSBsa2xqIGNxbnAgbWlvdQ==")      
        
        data_avui = datetime.now().strftime("%d-%m-%Y")
        self.var_data_inici = tk.StringVar(value="09-12-2025")
        self.var_data_fi = tk.StringVar(value=data_avui)

        tk.Label(root, text="Gestor de Faltes (V50)", font=("Arial", 16, "bold")).pack(pady=10)

        frame_email = tk.LabelFrame(root, text="Configuració Enviament", padx=10, pady=5, fg="blue")
        frame_email.pack(fill="x", padx=10, pady=5)
        tk.Label(frame_email, text="Email:").pack(side="left")
        tk.Entry(frame_email, textvariable=self.var_email_origen, width=35).pack(side="left", padx=5)
        tk.Label(frame_email, text="Pwd App:").pack(side="left")
        tk.Entry(frame_email, textvariable=self.var_password, width=20, show="*").pack(side="left", padx=5)

        frame_chrome = tk.LabelFrame(root, text="Configuració Chrome", padx=10, pady=5, fg="green")
        frame_chrome.pack(fill="x", padx=10, pady=5)
        tk.Label(frame_chrome, text="Versió de Chrome:").pack(side="left")
        tk.Entry(frame_chrome, textvariable=self.var_versio_chrome, width=10).pack(side="left", padx=5)

        frame_dates = tk.LabelFrame(root, text="Rang de Dates", padx=10, pady=5)
        frame_dates.pack(fill="x", padx=10, pady=5)
        tk.Label(frame_dates, text="Inici:").pack(side="left")
        tk.Entry(frame_dates, textvariable=self.var_data_inici, width=12).pack(side="left", padx=5)
        tk.Label(frame_dates, text="Fi:").pack(side="left", padx=10)
        tk.Entry(frame_dates, textvariable=self.var_data_fi, width=12).pack(side="left", padx=5)
        
        frame_carpeta = tk.LabelFrame(root, text="Carpeta de Treball", padx=10, pady=5)
        frame_carpeta.pack(fill="x", padx=10, pady=5)
        lbl_info = f"⚠️ Cal tenir:\n 1. '{NOM_FITXER_MESTRE_XLSX}' (Fulls: Alumnes, Tutors)\n 2. '{NOM_FITXER_EXPEDIENTS}' (Opcional, per filtrar)"
        tk.Label(frame_carpeta, text=lbl_info, fg="grey", justify="left").pack(anchor="w", padx=5)
        tk.Entry(frame_carpeta, textvariable=self.carpeta_desti, width=50).pack(side="left", padx=5)
        tk.Button(frame_carpeta, text="📂 Triar...", command=self.seleccionar_carpeta).pack(side="left")

        frame_accions = tk.Frame(root, pady=10)
        frame_accions.pack(fill="x", padx=10)
        tk.Button(frame_accions, text="1. OBRIR CHROME", bg="#d9edf7", command=self.iniciar_navegador).pack(side="left", fill="x", expand=True, padx=5)
        self.btn_baixar = tk.Button(frame_accions, text="2. BAIXAR AUTOMÀTIC", bg="#dff0d8", command=self.iniciar_proces, state="disabled")
        self.btn_baixar.pack(side="left", fill="x", expand=True, padx=5)
        tk.Button(frame_accions, text="3. ANALITZAR LOCAL", bg="#fcf8e3", command=self.analitzar_local).pack(side="left", fill="x", expand=True, padx=5)

        self.log_area = scrolledtext.ScrolledText(root, height=18)
        self.log_area.pack(fill="both", padx=10, pady=10, expand=True)
        
        # --- Peu de pàgina (Footer) ---
        footer_frame = tk.Frame(self.root, height=30, bg="#f0f0f0")
        footer_frame.pack(side="bottom", fill="x")
        
        llicencia_label = tk.Label(
            footer_frame,
            text="Llicència: CC BY-NC-SA 4.0",
            fg="blue",
            cursor="hand2",
            bg="#f0f0f0"
        )
        font_subratllada = font.Font(llicencia_label, llicencia_label.cget("font"))
        font_subratllada.configure(underline=True)
        llicencia_label.configure(font=font_subratllada)
        llicencia_label.bind("<Button-1>", self.obrir_llicencia)
        llicencia_label.pack(pady=5)
        
        print("   -> Interfície carregada correctament.")

    def log(self, missatge):
        self.log_area.insert(tk.END, f"• {missatge}\n")
        self.log_area.see(tk.END)
        self.root.update_idletasks()

    def seleccionar_carpeta(self):
        c = filedialog.askdirectory()
        if c: self.carpeta_desti.set(c)

    def obrir_llicencia(self, event):
        webbrowser.open_new(r"https://creativecommons.org/licenses/by-nc-sa/4.0/")

    def llegir_excel_universal(self, ruta, full=0):
        if ruta.endswith(".ods"):
            return pd.read_excel(ruta, engine="odf", sheet_name=full)
        else:
            return pd.read_excel(ruta, engine="openpyxl", sheet_name=full)

    def analitzar_local(self):
        carpeta_treball = self.carpeta_desti.get()
        if not carpeta_treball:
            messagebox.showwarning("Atenció", "Primer selecciona la 'Carpeta de Treball'.")
            return
        ruta_fitxer = filedialog.askopenfilename(title="Selecciona Faltes", filetypes=[("Excel", "*.xlsx *.ods")])
        if not ruta_fitxer: return
        d1 = self.var_data_inici.get().strip()
        d2 = self.var_data_fi.get().strip()
        threading.Thread(target=self.tractar_dades_excel, args=(ruta_fitxer, carpeta_treball, d1, d2), daemon=True).start()

    def iniciar_navegador(self):
        try:
            versio = int(self.var_versio_chrome.get())
            if os.path.exists(self.carpeta_temp): shutil.rmtree(self.carpeta_temp, ignore_errors=True)
            os.makedirs(self.carpeta_temp, exist_ok=True)
            threading.Thread(target=self._obrir_chrome_thread, args=(versio,), daemon=True).start()
        except ValueError: messagebox.showerror("Error", "Versió Chrome incorrecta")

    def _obrir_chrome_thread(self, versio):
        try:
            options = uc.ChromeOptions()
            prefs = {"download.default_directory": self.carpeta_temp, "safebrowsing.enabled": True}
            options.add_experimental_option("prefs", prefs)
            self.driver = uc.Chrome(options=options, use_subprocess=True, version_main=versio)
            self.driver.maximize_window()
            self.driver.get("https://insramonturroidarder.ieduca.com")
            self.root.after(0, lambda: self.log("✅ Chrome obert."))
            self.root.after(0, lambda: self.btn_baixar.config(state="normal"))
        except Exception as e:
            self.root.after(0, lambda e=e: self.log(f"❌ Error Chrome: {e}"))

    def iniciar_proces(self):
        if not self.carpeta_desti.get():
            messagebox.showwarning("Error", "Selecciona carpeta destí!")
            return
        threading.Thread(target=self.executar_tasca, daemon=True).start()

    def executar_tasca(self):
        self.root.after(0, lambda: self.btn_baixar.config(state="disabled"))
        d1 = self.var_data_inici.get().strip()
        d2 = self.var_data_fi.get().strip()
        
        url_final = f"{URL_BASE_FALTES}?data1={d1}&data2={d2}{PARAMS_FIXOS}"
        self.root.after(0, self.log, f"URL de descàrrega: {url_final}")
        
        nom_fitxer_faltes = f"Faltes_{d1}_a_{d2}.xlsx"
        carpeta_final = self.carpeta_desti.get()
        ok, msg = self.baixar_fitxer_directe(url_final, nom_fitxer_faltes, carpeta_final)
        if ok:
            ruta_completa = os.path.join(carpeta_final, nom_fitxer_faltes)
            self.root.after(0, self.log, f"✅ Dades baixades correctament a: {ruta_completa}")
            self.tractar_dades_excel(ruta_completa, carpeta_final, d1, d2)
        else:
            self.root.after(0, self.log, f"❌ Error descàrrega: {msg}")
        self.root.after(0, lambda: self.btn_baixar.config(state="normal"))

    def baixar_fitxer_directe(self, url, nom_desti, carpeta_desti):
        temps_inici = time.time()
        carpetes_a_buscar = [self.carpeta_temp, os.path.join(os.path.expanduser("~"), "Downloads")]
        self.root.after(0, self.log, f"🔎 Buscant el fitxer descarregat a: {', '.join(carpetes_a_buscar)}")

        try:
            self.root.after(0, self.log, "⏳ Sol·licitant fitxer al navegador...")
            self.driver.get(url)
            
            temps = 0
            while temps < 60:
                for carpeta in carpetes_a_buscar:
                    if not os.path.exists(carpeta): continue
                    
                    for f in os.listdir(carpeta):
                        if f.endswith('.xlsx') and not f.startswith("~"):
                            p = os.path.join(carpeta, f)
                            try:
                                time.sleep(1) 
                                if os.path.getmtime(p) > temps_inici and os.path.getsize(p) > 0:
                                    self.root.after(0, self.log, f"✔️ Fitxer trobat a: {p}")
                                    desti_final = os.path.join(carpeta_desti, nom_desti)
                                    shutil.move(p, desti_final)
                                    return True, "OK"
                            except FileNotFoundError:
                                time.sleep(0.5)
                                continue
                time.sleep(1)
                temps += 1
            return False, "Temps d'espera esgotat. No s'ha trobat el fitxer descarregat."
        except Exception as e:
            return False, str(e)

    def _categoritzar_fila(self, fila, col_nivell, col_tipus):
        text_tipus = ""
        if col_tipus and pd.notna(fila[col_tipus]):
            text_tipus = str(fila[col_tipus]).replace('.0', '').strip().upper()

        text_nivell = ""
        if col_nivell and pd.notna(fila[col_nivell]):
            text_nivell = str(fila[col_nivell]).replace('.0', '').strip().upper()

        if text_tipus == '4' or text_nivell == '4': return 'N4'
        if text_tipus == '3' or text_nivell == '3': return 'N3'
        if text_tipus == '2' or text_nivell == '2': return 'N2'
        
        if text_tipus == 'F': return 'INJUSTIFICADA'
        if text_tipus == 'M': return 'MATERIAL'
        if text_tipus == 'T': return 'TREBALL'

        if 'NIVELL 4' in text_tipus or 'MOLT GREU' in text_tipus: return 'N4'
        if 'NIVELL 3' in text_tipus or 'GREU' in text_tipus: return 'N3'
        if 'NIVELL 2' in text_tipus: return 'N2'
        
        if 'TREBALL' in text_tipus: return 'TREBALL'
        if 'INJUSTIFICADA' in text_tipus: return 'INJUSTIFICADA'
        if 'MATERIAL' in text_tipus: return 'MATERIAL'
        
        return None

    def normalitzar_nom(self, text):
        if not isinstance(text, str):
            return str(text).upper().strip()
        text = text.upper().strip()
        text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
        text = " ".join(text.split())
        return text

    def tractar_dades_excel(self, ruta_faltes, carpeta_treball, data_inici, data_fi):
        try:
            if not os.path.exists(ruta_faltes) or os.path.getsize(ruta_faltes) < 100:
                 self.root.after(0, self.log, f"⚠️ El fitxer de faltes '{os.path.basename(ruta_faltes)}' està buit o no existeix.")
                 self.root.after(0, messagebox.showwarning, "Fitxer Buit", "El fitxer de faltes descarregat està buit. No es pot continuar amb l'anàlisi.")
                 return

            ruta_alumnes = os.path.join(carpeta_treball, NOM_FITXER_MESTRE_XLSX)
            if not os.path.exists(ruta_alumnes):
                ruta_alumnes = os.path.join(carpeta_treball, NOM_FITXER_MESTRE_ODS)
            
            if not os.path.exists(ruta_alumnes):
                self.root.after(0, lambda: messagebox.showerror("Error", f"Falta '{NOM_FITXER_MESTRE_XLSX}' o '{NOM_FITXER_MESTRE_ODS}'"))
                return

            self.root.after(0, self.log, f"⚙️ Llegint mestre: {os.path.basename(ruta_alumnes)}")
            df_alumnes = self.llegir_excel_universal(ruta_alumnes, full="Alumnes")
            
            self.root.after(0, self.log, f"⚙️ Llegint fitxer de faltes: {os.path.basename(ruta_faltes)}")
            
            try:
                df_faltes = self.llegir_excel_universal(ruta_faltes)
            except Exception as e:
                self.root.after(0, self.log, f"❌ Error llegint fitxer de faltes '{os.path.basename(ruta_faltes)}': {e}")
                self.root.after(0, messagebox.showerror, "Error de Lectura", f"No s'ha pogut llegir el fitxer de faltes. Error: {e}")
                return # Retorna si hi ha error en la lectura

            if df_faltes.empty: # La comprovació de os.path.exists i os.path.getsize ja es fa a baixar_fitxer_directe
                 self.root.after(0, self.log, f"⚠️ El fitxer de faltes '{os.path.basename(ruta_faltes)}' no conté dades vàlides.")
                 self.root.after(0, messagebox.showwarning, "Fitxer Invàlid", "El fitxer de faltes descarregat no conté dades vàlides. No es pot continuar amb l'anàlisi.")
                 return
            
            mc_alu = {str(c).strip().lower(): c for c in df_alumnes.columns}
            
            def trobar_alu(p):
                for n in p: 
                    if n in mc_alu: return mc_alu[n]
                return None
            
            c_nom_sep = trobar_alu(['nom', 'nom alumne'])
            c_cog1_sep = trobar_alu(['cognom 1', 'cognom1', 'primer cognom'])
            c_cog2_sep = trobar_alu(['cognom 2', 'cognom2', 'segon cognom'])
            
            try:
                self.root.after(0, self.log, f"DIAGNÒSTIC: Intentant carregar la fulla 'Tutors' de '{os.path.basename(ruta_alumnes)}'.")
                df_tutors = self.llegir_excel_universal(ruta_alumnes, full="Tutors")
                self.root.after(0, self.log, f"DIAGNÒSTIC: Fulla 'Tutors' carregada amb {len(df_tutors)} files.")
            except Exception as e:
                self.root.after(0, self.log, f"DIAGNÒSTIC: No s'ha pogut carregar la fulla 'Tutors'. Error: {e}")
                df_tutors = pd.DataFrame()

            dict_expedients = {}
            ruta_exp = os.path.join(carpeta_treball, NOM_FITXER_EXPEDIENTS)
            if os.path.exists(ruta_exp):
                self.root.after(0, self.log, f"   -> Processant {NOM_FITXER_EXPEDIENTS}...")
                try:
                    df_exp = self.llegir_excel_universal(ruta_exp, full=NOM_FULL_EXPEDIENTS)
                    mc = {str(c).strip().lower(): c for c in df_exp.columns}
                    
                    def trobar(p):
                        for n in p: 
                            if n in mc: return mc[n]
                        return None
                    
                    c_c1 = trobar(['cognom 1', 'cognom1', 'primer cognom'])
                    c_c2 = trobar(['cognom 2', 'cognom2', 'segon cognom'])
                    c_nom = trobar(['nom'])
                    c_full_name = trobar(['nomcomplet', 'nom complet', 'alumne', 'alumne/a', 'estudiant', 'estudiant/a', 'estudiant_nom', 'nom_complet']) 
                    c_data_sanc = trobar(['data inici sanció', 'data sanció', 'data'])

                    if c_data_sanc:
                        for idx, row in df_exp.iterrows():
                            raw_name = None
                            if c_full_name and pd.notna(row[c_full_name]):
                                raw_name = row[c_full_name]
                            elif c_c1 and c_nom:
                                cog2 = row[c_c2] if c_c2 and pd.notna(row[c_c2]) else ''
                                if cog2:
                                    raw_name = f"{row[c_c1]} {cog2}, {row[c_nom]}"
                                else:
                                    raw_name = f"{row[c_c1]}, {row[c_nom]}"
                            
                            if raw_name:
                                nom_net = self.normalitzar_nom(raw_name)
                                dt = pd.to_datetime(row[c_data_sanc], dayfirst=True, errors='coerce')
                                if pd.notna(dt):
                                    if nom_net not in dict_expedients or dt > dict_expedients[nom_net]:
                                        dict_expedients[nom_net] = dt

                        self.root.after(0, self.log, f"   -> {len(dict_expedients)} expedients detectats.")
                    else:
                        self.root.after(0, self.log, f"⚠️ Error: No s'ha trobat la columna 'data inici sanció' o similar en {NOM_FITXER_EXPEDIENTS}.")
                except Exception as e:
                    self.root.after(0, self.log, f"⚠️ Error expedients: {str(e)}")
            
            llista_coordinadors = []
            if not df_tutors.empty:
                
                c_grup_tut = None
                c_email_tut = None
                
                # Find the group column by content
                for col_idx, col_name in enumerate(df_tutors.columns):
                    if df_tutors[col_name].astype(str).str.contains("COORDINADOR", case=False, na=False).any():
                        c_grup_tut = col_name
                        # Assume email column is the next one
                        if col_idx + 1 < len(df_tutors.columns):
                            c_email_tut = df_tutors.columns[col_idx + 1]
                        break

                self.root.after(0, self.log, f"DIAGNÒSTIC: Columnes detectades a 'Tutors': Grup='{c_grup_tut}', Email='{c_email_tut}'.")
                if c_grup_tut and c_email_tut:
                    for _, row in df_tutors.iterrows():
                        if "COORDINADOR" in str(row[c_grup_tut]).upper() and pd.notna(row[c_email_tut]) and "@" in str(row[c_email_tut]):
                            llista_coordinadors.append(row[c_email_tut])
                    self.root.after(0, self.log, f"DIAGNÒSTIC: S'han trobat {len(llista_coordinadors)} correus de coordinadors: {llista_coordinadors}.")
                else:
                    self.root.after(0, self.log, f"DIAGNÒSTIC: No s'han pogut identificar les columnes de grup o email a la fulla 'Tutors'.")


            mapa_cols_faltes = {str(c).strip().lower(): c for c in df_faltes.columns}
            mapa_cols_alumnes = {str(c).strip().lower(): c for c in df_alumnes.columns}
            
            def trobar_col(mapa, noms):
                for n in noms:
                    if n.lower() in mapa: return mapa[n.lower()]
                return None

            c_alumne_f = trobar_col(mapa_cols_faltes, ['alumne', 'nom', 'alumne/a'])
            c_alumne_a = trobar_col(mapa_cols_alumnes, ['nomcomplet', 'nom complet', 'alumne'])
            c_curs = trobar_col(mapa_cols_alumnes, ['curs 2025-2026', 'curs', 'grup'])
            c_pare = trobar_col(mapa_cols_alumnes, ['pare', 'tutor 1'])
            c_tel = trobar_col(mapa_cols_alumnes, ['tel1', 'telefon'])
            c_nivell = trobar_col(mapa_cols_faltes, ['nivell', 'gravetat'])
            c_tipus = trobar_col(mapa_cols_faltes, ['tipus', 'incidència'])
            c_data_falta = trobar_col(mapa_cols_faltes, ['data', 'dia'])

            # NOU: Comprovació de columnes essencials amb missatge detallat per a la data
            if not c_alumne_f or not c_alumne_a or not c_curs or not c_data_falta:
                missing = []
                if not c_alumne_f: missing.append("'alumne/a' (fitxer de faltes)")
                if not c_alumne_a: missing.append("'nomcomplet' (dades.xlsx)")
                if not c_curs: missing.append("'curs' (dades.xlsx)")
                if not c_data_falta: missing.append("'data' o 'dia' (fitxer de faltes)")
                
                msg = f"No s'han trobat les següents columnes essencials: {', '.join(missing)}.\n"
                msg += f"Columnes disponibles al fitxer de faltes: {list(df_faltes.columns)}."

                self.root.after(0, lambda: self.log(f"❌ Error crític: {msg}"))
                messagebox.showerror("Error de Columnes", msg)
                return

            df_faltes['__KEY'] = df_faltes[c_alumne_f].apply(self.normalitzar_nom)
            df_alumnes['__KEY'] = df_alumnes[c_alumne_a].apply(self.normalitzar_nom)
            df_faltes['__DATA_OBJ'] = pd.to_datetime(df_faltes[c_data_falta], dayfirst=True, errors='coerce')

            cols_a_merge = [c for c in [c_curs, c_pare, c_tel, c_nom_sep, c_cog1_sep, c_cog2_sep] if c]
            df_merged = pd.merge(df_faltes, df_alumnes[['__KEY'] + cols_a_merge], on='__KEY', how='left')
            df_merged[c_curs] = df_merged[c_curs].fillna("SENSE_CURS")

            self.generar_resum_global_i_enviar(df_merged, c_curs, c_alumne_f, c_pare, c_tel, c_tipus, c_nivell, c_data_falta,
                                               c_nom_sep, c_cog1_sep, c_cog2_sep,
                                               carpeta_treball, data_inici, data_fi, llista_coordinadors, dict_expedients)

        except Exception as e:
            self.root.after(0, self.log, f"❌ Error fatal en el tractament de dades: {e}")
            traceback.print_exc()
            
    def _aplicar_estils_full(self, ws, d1, d2):
        ws['A1'] = "RESUM GLOBAL INCIDÈNCIES"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"{d1} - {d2}"
        
        red = PatternFill("solid", fgColor="FFCCCC")
        yel = PatternFill("solid", fgColor="FFFFCC")
        
        amples = {'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 30, 'F': 5, 'G': 5, 'H': 5, 'I': 12, 'J': 12, 'K': 10, 'L': 25, 'M': 12}
        for col_letra, ample in amples.items():
            # Comprova si la columna existeix abans d'intentar accedir-hi
            if col_letra in ws.column_dimensions:
                 ws.column_dimensions[col_letra].width = ample

        header_map = {cell.value: cell.column_letter for cell in ws[4]}
        col_n3 = header_map.get("N3")
        col_n4 = header_map.get("N4")

        if col_n3 and col_n4:
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
                try:
                    n3_val = int(ws[f"{col_n3}{row[0].row}"].value)
                    n4_val = int(ws[f"{col_n4}{row[0].row}"].value)
                    
                    if n4_val > 0 or n3_val >= 3:
                        for c in row: c.fill = red
                    elif n3_val == 2:
                        for c in row: c.fill = yel
                except (ValueError, TypeError):
                    pass

    def generar_resum_global_i_enviar(self, df, c_curs, c_alum, c_pare, c_tel, c_tip, c_niv, c_data, 
                                      c_nom_sep, c_cog1_sep, c_cog2_sep, 
                                      carpeta, d1, d2, emails_coords, dict_expedients):
        self.root.after(0, self.log, "📊 Generant Resum Global...")
        self.root.after(0, self.log, f"DIAGNÒSTIC (generar_resum_global_i_enviar): Correus de coordinadors: {emails_coords}")


        if df.empty or c_alum not in df.columns:
            self.root.after(0, self.log, "✅ El dataframe de faltes està buit o és invàlid.")
            self.root.after(0, messagebox.showinfo, "Resum Buit", "No s'han trobat faltes per analitzar.")
            return
        
        res = []
        uniq = df[[c_curs, c_alum, '__KEY']].drop_duplicates().dropna(subset=[c_alum])
        
        for index, r in uniq.iterrows():
            sub = df[df[c_alum] == r[c_alum]].copy()
            nom_net_alum = self.normalitzar_nom(r[c_alum])
            data_exp_txt = ""
            if nom_net_alum in dict_expedients:
                limit = dict_expedients[nom_net_alum]
                sub = sub[sub['__DATA_OBJ'] >= limit]
                data_exp_txt = f" (Exp: {limit.strftime('%d/%m')})"

            if sub.empty: continue

            comptadors = {'N4': 0, 'N3': 0, 'N2': 0, 'TREBALL': 0, 'INJUSTIFICADA': 0, 'MATERIAL': 0}
            for idx_f, f in sub.iterrows():
                try:
                    cat = self._categoritzar_fila(f, c_niv, c_tip)
                    if cat in comptadors:
                        comptadors[cat] += 1
                except Exception as e:
                    self.root.after(0, self.log, f"❌ ERROR en categoritzar la fila amb índex {idx_f}. Error: {e}")
                    continue
            
            if comptadors['N4'] > 0 or comptadors['N3'] > 0:
                fila = sub.iloc[0]
                res.append({
                    '__KEY': r['__KEY'], # CORRECCIÓ: S'afegeix la clau única per enllaçar dades
                    "Curs": fila[c_curs],
                    "Cognom 1": fila[c_cog1_sep] if c_cog1_sep and pd.notna(fila[c_cog1_sep]) else '',
                    "Cognom 2": fila[c_cog2_sep] if c_cog2_sep and pd.notna(fila[c_cog2_sep]) else '',
                    "Nom": fila[c_nom_sep] if c_nom_sep and pd.notna(fila[c_nom_sep]) else '',
                    "Nom Complet": f"{r[c_alum]}{data_exp_txt}",
                    "N4": comptadors['N4'],
                    "N3": comptadors['N3'],
                    "N2": comptadors['N2'],
                    "F. Injustif.": comptadors['INJUSTIFICADA'],
                    "F. Material": comptadors['MATERIAL'],
                    "Treball": comptadors['TREBALL'],
                    "Pare/Mare": fila[c_pare] if c_pare and pd.notna(fila[c_pare]) else '-',
                    "Telèfon": str(fila[c_tel]).replace('.0', '') if c_tel and pd.notna(fila[c_tel]) else '-',
                })
        
        if not res:
            self.root.after(0, self.log, "✅ Resum buit. Cap alumne compleix els criteris (N3>0 o N4>0).")
            self.root.after(0, messagebox.showinfo, "Finalitzat", "Procés finalitzat. Cap alumne té N3 o N4 per aparèixer al resum.")
            return

        df_res = pd.DataFrame(res)
        
        df_res['s'] = df_res['Curs'].astype(str).apply(lambda x: (1 if "BAT" in x.upper() else (2 if "ESO" in x.upper() else 3), x))
        
        conditions = [
            (df_res['N4'] > 0) | (df_res['N3'] >= 3),
            (df_res['N3'] == 2),
            (df_res['N3'] == 1)
        ]
        choices = [1, 2, 3]
        df_res['priority'] = np.select(conditions, choices, default=4)
        
        df_res = df_res.sort_values(['s', 'priority']).drop(columns=['s', 'priority'])
        
        df_greus = df_res[(df_res['N4'] > 0) | (df_res['N3'] >= 3)].copy()

        ruta_excel = os.path.join(carpeta, "RESUM_GLOBAL.xlsx")

        with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
            # Drop la clau interna abans de desar per no mostrar-la a l'usuari
            df_res.drop(columns=['__KEY']).to_excel(writer, sheet_name='Resum General', index=False, startrow=3)
            if not df_greus.empty:
                df_greus.drop(columns=['__KEY']).to_excel(writer, sheet_name='Casos Greus', index=False, startrow=3)

        wb = load_workbook(ruta_excel)
        self._aplicar_estils_full(wb['Resum General'], d1, d2)
        if 'Casos Greus' in wb.sheetnames:
            self._aplicar_estils_full(wb['Casos Greus'], d1, d2)
        wb.save(ruta_excel)

        if not df_greus.empty:
            self.generar_informes_md(df, df_greus, c_curs, c_alum, c_tip, c_niv, c_data, carpeta, dict_expedients)

        if emails_coords:
            self.root.after(0, self.log, "📧 Enviant resum...")
            assumpte = f"RESUM GLOBAL Faltes ({d1} - {d2})"
            cos = f"""Bon dia,

Us adjunto el resum global de faltes d'assistència corresponent al període del {d1} al {d2}.

Aquest document inclou:
- Un resum general amb tots els alumnes que tenen incidències de nivell 3 o 4.
- Una pestanya específica de 'Casos Greus' per a un seguiment més acurat.

Si us plau, feu una valoració d'aquestes dades amb els tutors corresponents dels vostres grups.

**Nota important:** Pel que fa a la informació dels Cicles Formatius (CF), tingueu en compte que les dades poden no ser totalment exactes.

Salutacions cordials,

Cap d'Estudis"""
            adjunts = [ruta_excel]
            for mail in emails_coords: self.enviar_correu(mail, assumpte, cos, adjunts)
        self.root.after(0, messagebox.showinfo, "Fet", "Procés finalitzat!")

    def generar_informes_md(self, df_total, df_greus, col_curs, col_alumne, c_tipus, c_nivell, c_data, carpeta_sortida, dict_expedients):
        self.root.after(0, lambda: self.log(f"📝 Generant informes MD millorats..."))
        
        try:
            carpeta_informes = os.path.join(carpeta_sortida, "informes")
            os.makedirs(carpeta_informes, exist_ok=True)

            c_profe = next((c for c in df_total.columns if 'professor' in str(c).lower()), None)
            c_obs = next((c for c in df_total.columns if 'observaci' in str(c).lower() or 'descripci' in str(c).lower()), None)
            
            cursos = df_greus["Curs"].unique()

            for curs in cursos:
                nom_net_curs = str(curs).replace("/", "-").strip()
                ruta_md = os.path.join(carpeta_informes, f"Informe_{nom_net_curs}.md")
                
                df_curs_greus = df_greus[df_greus["Curs"] == curs]

                with open(ruta_md, "w", encoding="utf-8") as f:
                    f.write(f"# 📄 Informe d'Assistència: {curs}\n")
                    f.write(f"*{datetime.now().strftime('%d/%m/%Y')}*\n\n")
                    f.write("---\n")

                    for _, alumne_row in df_curs_greus.iterrows():
                        alumne_key = alumne_row['__KEY']
                        alumne_nom_complet = alumne_row['Nom Complet']
                        
                        pare = alumne_row.get('Pare/Mare', '-')
                        tel = str(alumne_row.get('Telèfon', '-')).replace('.0', '')

                        f.write(f"\n## 👤 {alumne_nom_complet}\n")
                        f.write(f"- 📞 **Contacte:** {pare} / {tel}\n\n")

                        df_al_complet = df_total[df_total['__KEY'] == alumne_key].copy()
                        
                        nom_net_alum = self.normalitzar_nom(alumne_nom_complet.split(' (Exp:')[0])
                        if nom_net_alum in dict_expedients:
                            data_limit = dict_expedients[nom_net_alum]
                            self.root.after(0, self.log, f"DEBUG (MD): Alumne: '{nom_net_alum}', expedient registrat amb data_limit: {data_limit.strftime('%d/%m/%Y')}")
                            if '__DATA_OBJ' in df_al_complet.columns:
                                df_al_complet = df_al_complet[df_al_complet['__DATA_OBJ'] >= data_limit]
                        
                        if df_al_complet.empty:
                            f.write("*Sense incidències recents per a aquest alumne.*\n")
                            f.write("\n---\n")
                            continue

                        df_al_complet['__CATEGORIA'] = df_al_complet.apply(
                            lambda row: self._categoritzar_fila(row, c_nivell, c_tipus), axis=1
                        )
                        
                        df_n4 = df_al_complet[df_al_complet['__CATEGORIA'] == 'N4'].sort_values('__DATA_OBJ', ascending=True)
                        df_n3 = df_al_complet[df_al_complet['__CATEGORIA'] == 'N3'].sort_values('__DATA_OBJ', ascending=True)

                        if df_n4.empty and df_n3.empty:
                            f.write("*Sense faltes de Nivell 3 o 4 en el període seleccionat.*\n")
                        else:
                            if not df_n4.empty:
                                f.write("### 🔴 Faltes de Nivell 4\n\n")
                                for i, (_, falta) in enumerate(df_n4.iterrows()):
                                    if i > 0: f.write("- ---\n")
                                    data_str = falta['__DATA_OBJ'].strftime('%d/%m/%Y') if pd.notna(falta['__DATA_OBJ']) else "?"
                                    profe_str = falta[c_profe] if c_profe and pd.notna(falta[c_profe]) else "N/D"
                                    obs_str = falta[c_obs] if c_obs and pd.notna(falta[c_obs]) and str(falta[c_obs]).strip() != "" else ""
                                    f.write(f"- **Data:** {data_str}\n")
                                    f.write(f"  - **Professor/a:** {profe_str}\n")
                                    if obs_str:
                                        f.write(f"  - **Observacions:** {obs_str}\n\n")
                                f.write("\n")

                            if not df_n3.empty:
                                f.write("### 🟠 Faltes de Nivell 3\n\n")
                                for i, (_, falta) in enumerate(df_n3.iterrows()):
                                    if i > 0: f.write("- ---\n")
                                    data_str = falta['__DATA_OBJ'].strftime('%d/%m/%Y') if pd.notna(falta['__DATA_OBJ']) else "?"
                                    profe_str = falta[c_profe] if c_profe and pd.notna(falta[c_profe]) else "N/D"
                                    obs_str = falta[c_obs] if c_obs and pd.notna(falta[c_obs]) and str(falta[c_obs]).strip() != "" else ""
                                    f.write(f"- **Data:** {data_str}\n")
                                    f.write(f"  - **Professor/a:** {profe_str}\n")
                                    if obs_str:
                                        f.write(f"  - **Observacions:** {obs_str}\n\n")
                        
                        f.write("\n---\n") # Final de la secció de l'alumne
            self.root.after(0, lambda: self.log(f"✅ Informes MD (millorats) generats a la carpeta 'informes'."))
        except Exception as e:
            print("\n" + "*"*20 + " ERROR EN GENERAR INFORMES MD " + "*"*20)
            print(f"S'ha produït un error inesperat en la nova funció de generar MD: {e}")
            traceback.print_exc()
            print("*"*65 + "\n")
            self.root.after(0, lambda err=e: messagebox.showerror("Error Informes MD", f"S'ha produït un error en generar els informes MD. Revisa la terminal per a més detalls. Error: {err}"))

    def enviar_correu(self, destinatari, assumpte, cos, fitxers_adjunts=[]):
        remitent = self.var_email_origen.get().strip()
        password = base64.b64decode(self.var_password.get().replace(" ", "").strip()).decode('utf-8')
        if not remitent or not password: return
        msg = MIMEMultipart(); msg['From'] = remitent; msg['To'] = destinatari; msg['Subject'] = assumpte
        msg.attach(MIMEText(cos, 'plain'))
        for f in fitxers_adjunts:
            if os.path.exists(f):
                with open(f, "rb") as attachment:
                    part = MIMEApplication(attachment.read(), Name=os.path.basename(f))
                    part['Content-Disposition'] = f'attachment; filename="{os.path.basename(f)}"'
                    msg.attach(part)
        try:
            server = smtplib.SMTP('smtp.gmail.com', 587); server.starttls(); server.login(remitent, password)
            server.send_message(msg); server.quit()
            self.root.after(0, self.log, f"   -> ✉️ Enviat a: {destinatari}")
        except Exception as e:
            self.root.after(0, self.log, f"❌ Error enviament: {e}")

if __name__ == "__main__":
    print("⏳ [5/5] Llançant interfície gràfica...")
    root = tk.Tk()
    app = AplicacioFaltes(root)
    print("✅ Bucle principal iniciat.")
    root.mainloop()