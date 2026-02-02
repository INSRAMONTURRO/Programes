# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------
# Copia_Format_fulls_v5.py
#
# Script per processar fitxers Excel d'actes, duplicant una fulla
# i aplicant formats i imatges.
#
# Autor: Josep Maria Sardà
# Versió: 5.0
# Data: 02 de Febrer de 2026
# Llicència: Creative Commons BY-NC-SA 4.0
#
# Canvis:
# - v5.0: Afegida autoria i llicència al codi.
# ----------------------------------------------------------------------

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font
import webbrowser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.print_settings import PrintArea
import pandas as pd
from copy import copy

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador d'Excel - Actes amb imatges (Millorada)")
        self.root.geometry("750x500")

        # Variables
        self.source_folder = tk.StringVar()
        self.destination_folder = tk.StringVar()
        self.file_prefix = tk.StringVar(value="1aAVA")
        self.image_path = tk.StringVar(value="")  # Variable per emmagatzemar la ruta de la imatge

        # Noms from CSV
        self.noms = []
        self.load_noms_from_csv()

        self.setup_ui()

    def load_noms_from_csv(self):
        """Carrega els noms del fitxer CSV si existeix"""
        try:
            with open("Noms.csv", "r", encoding="utf-8") as f:
                self.noms = [line.strip() for line in f.readlines()]
        except FileNotFoundError:
            messagebox.showwarning("Advertència", "No s'ha trobat el fitxer Noms.csv. S'utilitzarà una llista per defecte.")
            self.noms = ["1 ESO A", "1 ESO B", "1 ESO C", "2 ESO A", "2 ESO B", "2 ESO C"]

    def setup_ui(self):
        """Configura la interfície d'usuari"""
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Títol
        title_label = ttk.Label(main_frame, text="Processador d'Excel d'Actes amb Imatges", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

        # Selecció carpeta origen
        ttk.Label(main_frame, text="Carpeta Origen:").grid(row=1, column=0, sticky=tk.W, pady=5)
        origin_frame = ttk.Frame(main_frame)
        origin_frame.grid(row=1, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        ttk.Entry(origin_frame, textvariable=self.source_folder, width=50).grid(row=0, column=0, sticky=(tk.W, tk.E))
        ttk.Button(origin_frame, text="...", command=self.select_source_folder).grid(row=0, column=1, padx=(5, 0))
        origin_frame.columnconfigure(0, weight=1)

        # Selecció carpeta destí
        ttk.Label(main_frame, text="Carpeta Destí:").grid(row=2, column=0, sticky=tk.W, pady=5)
        dest_frame = ttk.Frame(main_frame)
        dest_frame.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        ttk.Entry(dest_frame, textvariable=self.destination_folder, width=50).grid(row=0, column=0, sticky=(tk.W, tk.E))
        ttk.Button(dest_frame, text="...", command=self.select_destination_folder).grid(row=0, column=1, padx=(5, 0))
        dest_frame.columnconfigure(0, weight=1)

        # Selecció imatge
        ttk.Label(main_frame, text="Imatge a afegir:").grid(row=3, column=0, sticky=tk.W, pady=5)
        img_frame = ttk.Frame(main_frame)
        img_frame.grid(row=3, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        ttk.Entry(img_frame, textvariable=self.image_path, width=50).grid(row=0, column=0, sticky=(tk.W, tk.E))
        ttk.Button(img_frame, text="...", command=self.select_image).grid(row=0, column=1, padx=(5, 0))
        img_frame.columnconfigure(0, weight=1)

        # Prefix del nom dels fitxers
        ttk.Label(main_frame, text="Prefix del nom:").grid(row=4, column=0, sticky=tk.W, pady=5)
        prefix_combo = ttk.Combobox(main_frame, textvariable=self.file_prefix, values=["1aAVA", "2aAVA", "3aAVA", "Final", "Extraor"], state="readonly")
        prefix_combo.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5)

        # Llistat de fitxers a processar
        ttk.Label(main_frame, text="Fitxers a processar:").grid(row=5, column=0, sticky=(tk.W, tk.N), pady=5)
        self.files_listbox = tk.Listbox(main_frame, height=8)
        self.files_listbox.grid(row=5, column=1, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

        # Botó de processament
        self.process_button = ttk.Button(main_frame, text="Processa Fitxers", command=self.process_files)
        self.process_button.grid(row=6, column=0, columnspan=3, pady=20)

        # Barra de progrés
        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)

        # Missatge d'estat
        self.status_label = ttk.Label(main_frame, text="Si us plau, seleccioni les carpetes origen i destí")
        self.status_label.grid(row=8, column=0, columnspan=3, pady=5)

        # Configuració de la grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # --- Peu de pàgina (Footer) ---
        footer_frame = ttk.Frame(self.root, style="Footer.TFrame")
        footer_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))

        style = ttk.Style()
        style.configure("Footer.TFrame", background="#f0f0f0")

        autor = "Josep Maria Sardà"
        llicencia_text = "CC BY-NC-SA 4.0"
        url_llicencia = "https://creativecommons.org/licenses/by-nc-sa/4.0/"

        info_text = f"Autor: {autor}  |  Llicència: {llicencia_text}"

        footer_label = ttk.Label(
            footer_frame,
            text=info_text,
            foreground="blue",
            cursor="hand2",
            style="Footer.TLabel"
        )
        style.configure("Footer.TLabel", background="#f0f0f0")

        # Subratllar el text per semblar un enllaç
        font_subratllada = font.Font(footer_label, footer_label.cget("font"))
        font_subratllada.configure(underline=True)
        footer_label.configure(font=font_subratllada)

        # Assignar l'esdeveniment de clic
        footer_label.bind("<Button-1>", self.obrir_llicencia)
        footer_label.pack(pady=5)
        
        self.root.rowconfigure(1, weight=0)

    def obrir_llicencia(self, event):
        webbrowser.open_new(r"https://creativecommons.org/licenses/by-nc-sa/4.0/")

    def select_image(self):
        """Selecciona la imatge a afegir"""
        file_path = filedialog.askopenfilename(
            title="Selecciona la imatge",
            filetypes=[("Imatges", "*.png *.jpg *.jpeg *.gif *.bmp *.tiff *.webp")]
        )
        if file_path:
            self.image_path.set(file_path)

    def select_source_folder(self):
        """Selecciona la carpeta origen"""
        folder = filedialog.askdirectory(title="Selecciona la carpeta origen")
        if folder:
            self.source_folder.set(folder)
            self.update_files_list()

    def select_destination_folder(self):
        """Selecciona la carpeta destí"""
        folder = filedialog.askdirectory(title="Selecciona la carpeta destí")
        if folder:
            self.destination_folder.set(folder)

    def update_files_list(self):
        """Actualitza la llista de fitxers a processar"""
        self.files_listbox.delete(0, tk.END)

        if not self.source_folder.get():
            return

        source_path = self.source_folder.get()
        excel_files = [f for f in os.listdir(source_path) if f.endswith('.xlsx') and not f.startswith('~$')]

        for file in excel_files:
            self.files_listbox.insert(tk.END, file)

    def process_files(self):
        """Processa els fitxers Excel"""
        if not self.source_folder.get() or not self.destination_folder.get():
            messagebox.showerror("Error", "Si us plau, seleccioni les carpetes origen i destí")
            return

        source_path = self.source_folder.get()
        dest_path = self.destination_folder.get()

        # Obté els fitxers a processar
        selected_files = [self.files_listbox.get(i) for i in range(self.files_listbox.size())]

        if not selected_files:
            messagebox.showwarning("Advertència", "No hi ha fitxers per processar")
            return

        self.progress['maximum'] = len(selected_files)
        self.status_label.config(text="Processant fitxers...")
        self.process_button.config(state='disabled')

        errors_occurred = []
        try:
            for i, filename in enumerate(selected_files):
                source_file = os.path.join(source_path, filename)

                # Determina el nom del destí basat en CSV
                dest_filename = self.get_destination_filename(filename)
                dest_file = os.path.join(dest_path, dest_filename)

                # Processa el fitxer
                try:
                    self.process_single_file(source_file, dest_file)
                except Exception as e:
                    error_msg = f"Error processant {filename}: {str(e)}"
                    errors_occurred.append(error_msg)
                    print(error_msg)  # Log to console

                # Actualitza la barra de progrés
                self.progress['value'] = i + 1
                self.root.update_idletasks()

            if errors_occurred:
                error_text = f"Processament completat amb {len(errors_occurred)} errors:\n" + "\n".join(errors_occurred[:3])  # Show first 3 errors
                if len(errors_occurred) > 3:
                    error_text += f"\n... i {len(errors_occurred) - 3} més"
                messagebox.showwarning("Completat amb errors", error_text)
            else:
                self.status_label.config(text=f"Processament completat! {len(selected_files)} fitxers processats.")
                messagebox.showinfo("Completat", f"S'han processat {len(selected_files)} fitxers correctament!")

        except Exception as e:
            messagebox.showerror("Error", f"S'ha produït un error: {str(e)}")

        finally:
            self.process_button.config(state='normal')
            self.progress['value'] = 0

    def get_destination_filename(self, original_filename):
        """Obté el nom del fitxer de destinació basat en el CSV"""
        # Extrau la part del nom que coincideix amb un dels valors del CSV
        original_name = os.path.splitext(original_filename)[0]

        # Cerca coincidències amb els noms del CSV
        matching_name = None
        for nom in self.noms:
            if nom.replace(" ", "") in original_name.replace(" ", ""):  # Comprovació per nom sense espais
                matching_name = nom.replace(" ", "_").replace("/", "_")  # Substitueix espais i barres per guions baixos
                break

        # Si no troba coincidència, utilitza el nom original
        if not matching_name:
            matching_name = original_name.replace(" ", "_")

        # Construeix el nom del fitxer de destinació
        # Independentment del prefix seleccionat, el nom haurà de començar amb "Acta_"
        # seguit del prefix i després el nom que coincideix
        if self.file_prefix.get() == "Acta":
            return f"Acta_1AVA_{matching_name}.xlsx"
        else:
            # Per qualsevol altre prefix (2aAVA, 3aAVA, etc.), el nom ha de començar amb "Acta_"
            return f"Acta_{self.file_prefix.get()}_{matching_name}.xlsx"

    def process_single_file(self, source_file, dest_file):
        """Processa un únic fitxer Excel"""
        from openpyxl.drawing.image import Image

        # Carrega el fitxer Excel
        wb = openpyxl.load_workbook(source_file)

        # Verifica si la fulla "1aAVA" existeix
        if "1aAVA" not in wb.sheetnames:
            # Si no existeix, utilitza la primera fulla
            source_sheet_name = wb.sheetnames[0]
        else:
            source_sheet_name = "1aAVA"

        source_sheet = wb[source_sheet_name]

        # Aplica configuració d'impressió a la fulla original també
        self.apply_print_settings(source_sheet)

        # Crea les noves fulles - copiant cel·les i estils
        new_sheet_names = ["2aAVA", "3aAVA", "Final", "Extraor"]

        for new_sheet_name in new_sheet_names:
            # Còpia la fulla origen (això copia cel·les i estils!)
            new_sheet = wb.copy_worksheet(source_sheet)
            new_sheet.title = new_sheet_name

            # Aplica configuració d'impressió a la nova fulla
            self.apply_print_settings(new_sheet)

        # Ara afegim la imatge a totes les fulles (incloent-hi la original)
        # per assegurar-nos que cada fulla només té una imatge
        if self.image_path.get() and os.path.exists(self.image_path.get()):
            all_sheet_names = [source_sheet_name] + new_sheet_names
            for sheet_name in all_sheet_names:
                sheet = wb[sheet_name]

                # Eliminar imatges existents en aquesta fulla (si n'hi hagués de copiar)
                sheet._images.clear()

                try:
                    img = Image(self.image_path.get())

                    # Ajustar la mida de la imatge a 1,84 cm d'amplada i 1,19 cm d'alçada
                    # openpyxl treballa amb píxels, cal convertir de cm a píxels
                    # Aproximadament: 1 cm ≈ 37.8 píxels a 96 DPI
                    img.width = 1.84 * 37.8  # Aproximadament 69.5 píxels
                    img.height = 1.19 * 37.8  # Aproximadament 44.9 píxels

                    # Col·loca la imatge a la fulla
                    sheet.add_image(img)

                except Exception as e:
                    print(f"Avís: No s'ha pogut afegir la imatge a la fulla {sheet_name}: {e}")

        # Desa el fitxer a la destinació
        wb.save(dest_file)

    def copy_cells_and_styles(self, source_sheet, dest_sheet):
        """Copia cel·les i estils d'una fulla a una altra"""
        for row in source_sheet.iter_rows():
            for cell in row:
                dest_cell = dest_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                # Copia estils
                if cell.has_style:
                    # Crear còpies explícites dels estils per evitar errors de desfasament
                    dest_cell.font = copy(cell.font)
                    dest_cell.border = copy(cell.border)
                    dest_cell.fill = copy(cell.fill)
                    dest_cell.number_format = cell.number_format
                    dest_cell.protection = copy(cell.protection)
                    dest_cell.alignment = copy(cell.alignment)

        # Copia amplades de columnes
        for col in source_sheet.column_dimensions:
            dest_col = dest_sheet.column_dimensions[col]
            source_col = source_sheet.column_dimensions[col]
            dest_col.width = source_col.width
            dest_col.hidden = source_col.hidden

        # Copia alçades de files
        for row in source_sheet.row_dimensions:
            dest_row = dest_sheet.row_dimensions[row]
            source_row = source_sheet.row_dimensions[row]
            dest_row.height = source_row.height
            dest_row.hidden = source_row.hidden

    def copy_images(self, source_sheet, dest_sheet):
        """Aquest mètode ja no s'utilitza en aquesta versió millorada"""
        pass

    def apply_print_settings(self, sheet):
        """Aplica la configuració d'impressió a la fulla"""
        # Força el mode d'ajust (evita conflictes entre scale i fitToPage)
        if sheet.sheet_properties.pageSetUpPr is None:
            from openpyxl.worksheet.properties import WorksheetProperties
            from openpyxl.worksheet.page_setup import PageSetup
            sheet.sheet_properties.pageSetUpPr = WorksheetProperties()
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        ps = sheet.page_setup
        ps.orientation = "portrait"
        ps.paperSize = 9  # equivalent a A4
        ps.fitToWidth = 1
        ps.fitToHeight = 2
        ps.scale = None  # Important: si uses fitTo, scale ha de ser None

        pm = sheet.page_margins
        pm.left = 0.39375
        pm.right = 0.39375
        pm.top = 0.39375
        pm.bottom = 0.39375
        pm.header = 0.511811023622047
        pm.footer = 0.511811023622047

        po = sheet.print_options
        po.gridLines = False
        po.headings = False
        po.horizontalCentered = False
        po.verticalCentered = False

def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
