# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------
# AnonimitzadorExcel.py
#
# Interfície gràfica (GUI) en Tkinter per a l'anonimització i desanonimització
# de columnes específiques de fitxers Excel de notes/visites, mantenint un
# mapeig consistent a un fitxer JSON.
#
# Autor: Josepm
# Versió: 2.2 (GUI Actualitzada - Multi-full, Prefixos i Normalització de Noms)
# Data: 28 de juny de 2026
# Llicència: Creative Commons BY-NC-SA 4.0
#
# Canvis:
# - v2.2: Normalització de noms (accents, majúscules/minúscules, espais consecutius) per evitar duplicats.
# - v2.1: S'ha estès el processament a TOTS els fulls del llibre Excel.
#         S'ha afegit un nou camp per configurar prefixos individuals per a
#         cada columna (ex: Professor per a la col. B, Alumne per a la C).
# - v2.0: Actualitzada la selecció per a fitxer Excel individual (en comptes de carpeta),
#         permet especificar múltiples columnes separades per comes (ex: B, C),
#         implementat processament asíncron amb Queue per a estabilitat,
#         i s'ha integrat la capçalera i peu de pàgina d'autoria del projecte.
# - v1.1: Versió original.
# ----------------------------------------------------------------------

import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog, ttk, font
import os
import json
import openpyxl
from openpyxl.utils import column_index_from_string
import threading
import queue
import webbrowser
from datetime import datetime
import re
import unicodedata

# --- CONFIGURACIÓ ---
FITXER_MAPPEIG = "mapeig_anonim.json"


def create_modern_button(parent, text, command, bg="#2563eb", fg="white", active_bg="#1d4ed8", font_size=10, bold=True):
    """Crea un botó modern i pla amb efecte d'hover."""
    weight = "bold" if bold else "normal"
    btn = tk.Button(
        parent,
        text=text,
        command=command,
        bg=bg,
        fg=fg,
        activebackground=active_bg,
        activeforeground=fg,
        font=("Helvetica", font_size, weight),
        relief="flat",
        bd=0,
        padx=12,
        pady=6,
        cursor="hand2"
    )
    def on_enter(e):
        btn.config(bg=active_bg)
    def on_leave(e):
        btn.config(bg=bg)
        
    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    return btn


class AnonimitzadorExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor d'Anonimització iEduca (v2.2)")
        self.root.geometry("750x800")
        self.root.minsize(600, 700)
        self.root.configure(bg="#f1f5f9")
        
        # Variables d'estat
        self.excel_file_path = tk.StringVar()
        self.columna_noms = tk.StringVar(value="B, D")
        self.prefixos_var = tk.StringVar(value="Professor, Alumne")
        self.paraula_clau = tk.StringVar(value="mitjanes")
        self.mode_reves = tk.BooleanVar(value=False) # False = Anonimitzar, True = Desanonimitzar
        
        self.mapeig = self.carregar_mapeig()
        
        # Cua per a fils de fons
        self.gui_queue = queue.Queue()
        self.check_queue()
        
        self.setup_ui()

    def setup_ui(self):
        # 1. Capçalera
        header_frame = tk.Frame(self.root, bg="#1e293b", height=75)
        header_frame.pack(side="top", fill="x")
        header_frame.pack_propagate(False)
        
        lbl_title = tk.Label(
            header_frame, 
            text="Anonimitzador de Fitxers Excel", 
            font=("Helvetica", 16, "bold"), 
            bg="#1e293b", 
            fg="#ffffff"
        )
        lbl_title.pack(anchor="w", padx=20, pady=(12, 0))
        
        lbl_sub = tk.Label(
            header_frame, 
            text="Anonimitza o desanonimitza columnes a tots els fulls d'un fitxer Excel", 
            font=("Helvetica", 9, "italic"), 
            bg="#1e293b", 
            fg="#94a3b8"
        )
        lbl_sub.pack(anchor="w", padx=20, pady=(0, 12))

        # Contenidor principal
        main_frame = tk.Frame(self.root, bg="#f1f5f9")
        main_frame.pack(side="top", fill="both", expand=True, padx=20, pady=15)

        # Card 0: Mode d'operació
        frame_mode = tk.LabelFrame(
            main_frame, 
            text=" Mode d'operació ", 
            font=("Helvetica", 10, "bold"), 
            bg="#f1f5f9", 
            fg="#1e293b", 
            bd=1, 
            relief="solid",
            padx=12, 
            pady=10
        )
        frame_mode.pack(fill="x", pady=(0, 15))
        
        tk.Radiobutton(
            frame_mode, 
            text="Anonimitzar (Noms Reals ➔ Pseudònim)", 
            variable=self.mode_reves, 
            value=False, 
            font=("Helvetica", 10, "bold"),
            bg="#f1f5f9",
            activebackground="#f1f5f9",
            fg="#2563eb",
            selectcolor="#ffffff"
        )        .pack(side="left", padx=15, pady=5)
        
        tk.Radiobutton(
            frame_mode, 
            text="Desanonimitzar (Pseudònim ➔ Noms Reals)", 
            variable=self.mode_reves, 
            value=True, 
            font=("Helvetica", 10, "bold"),
            bg="#f1f5f9",
            activebackground="#f1f5f9",
            fg="#dc2626",
            selectcolor="#ffffff"
        )        .pack(side="left", padx=15, pady=5)

        # Card 1: Selecció de Fitxer
        file_card = tk.LabelFrame(
            main_frame, 
            text=" 1. Selecciona el fitxer Excel a processar ", 
            font=("Helvetica", 10, "bold"), 
            bg="#f1f5f9", 
            fg="#1e293b", 
            bd=1, 
            relief="solid", 
            padx=12, 
            pady=10
        )
        file_card.pack(fill="x", pady=(0, 15))
        file_card.columnconfigure(0, weight=1)
        
        self.entry_file = tk.Entry(
            file_card, 
            textvariable=self.excel_file_path, 
            font=("Helvetica", 10), 
            state="readonly", 
            readonlybackground="#ffffff", 
            fg="#334155", 
            bd=1, 
            relief="solid"
        )
        self.entry_file.grid(row=0, column=0, sticky="ew", padx=(0, 10), ipady=5)
        
        self.btn_browse = create_modern_button(
            file_card, 
            text="Tria fitxer...", 
            command=self.seleccionar_fitxer,
            bg="#2563eb",
            fg="white",
            active_bg="#1d4ed8"
        )
        self.btn_browse.grid(row=0, column=1, sticky="e")

        # Card 2: Configuració de l'Anonimització
        config_card = tk.LabelFrame(
            main_frame, 
            text=" 2. Paràmetres d'anonimització ", 
            font=("Helvetica", 10, "bold"), 
            bg="#f1f5f9", 
            fg="#1e293b", 
            bd=1, 
            relief="solid", 
            padx=12, 
            pady=10
        )
        config_card.pack(fill="x", pady=(0, 15))
        config_card.columnconfigure(1, weight=1)
        
        lbl_cols = tk.Label(
            config_card, 
            text="Columnes a anonimitzar\n(separades per comes, ex: B, C):", 
            font=("Helvetica", 9, "bold"), 
            bg="#f1f5f9", 
            fg="#334155",
            justify="left"
        )
        lbl_cols.grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)
        
        entry_cols = tk.Entry(
            config_card, 
            textvariable=self.columna_noms, 
            font=("Helvetica", 10), 
            bd=1, 
            relief="solid"
        )
        entry_cols.grid(row=0, column=1, sticky="ew", ipady=5, pady=6)
        
        lbl_prefixes = tk.Label(
            config_card, 
            text="Prefixos de cada columna\n(separats per comes, ex: Professor, Alumne):", 
            font=("Helvetica", 9, "bold"), 
            bg="#f1f5f9", 
            fg="#334155",
            justify="left"
        )
        lbl_prefixes.grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        
        entry_prefixes = tk.Entry(
            config_card, 
            textvariable=self.prefixos_var, 
            font=("Helvetica", 10), 
            bd=1, 
            relief="solid"
        )
        entry_prefixes.grid(row=1, column=1, sticky="ew", ipady=5, pady=6)
        
        lbl_paraula = tk.Label(
            config_card, 
            text="Fila límit (paraula clau):\n(ex: mitjanes - opcional)", 
            font=("Helvetica", 9, "bold"), 
            bg="#f1f5f9", 
            fg="#334155",
            justify="left"
        )
        lbl_paraula.grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        
        entry_paraula = tk.Entry(
            config_card, 
            textvariable=self.paraula_clau, 
            font=("Helvetica", 10), 
            bd=1, 
            relief="solid"
        )
        entry_paraula.grid(row=2, column=1, sticky="ew", ipady=5, pady=6)
        
        lbl_hint = tk.Label(
            config_card, 
            text="Nota: En anonimitzar, s'esborraran les files que estiguin per sota de la paraula clau (a tots els fulls on es trobi).", 
            font=("Helvetica", 8, "italic"), 
            bg="#f1f5f9", 
            fg="#64748b"
        )
        lbl_hint.grid(row=3, column=0, columnspan=2, sticky="w", pady=(5, 0))

        # Card 3: Accions i Registre
        actions_card = tk.LabelFrame(
            main_frame, 
            text=" 3. Accions i Registre de progrés ", 
            font=("Helvetica", 10, "bold"), 
            bg="#f1f5f9", 
            fg="#1e293b", 
            bd=1, 
            relief="solid", 
            padx=12, 
            pady=10
        )
        actions_card.pack(fill="both", expand=True)
        
        buttons_frame = tk.Frame(actions_card, bg="#f1f5f9")
        buttons_frame.pack(fill="x", pady=(0, 10))
        
        self.btn_run = create_modern_button(
            buttons_frame, 
            text="EXECUTAR PROCESSAMENT A TOTS ELS FULLS", 
            command=self.start_processing,
            bg="#16a34a",
            fg="white",
            active_bg="#15803d",
            font_size=11,
            bold=True
        )
        self.btn_run.pack(side="left", fill="x", expand=True, padx=(0, 6))
        
        self.btn_reset = create_modern_button(
            buttons_frame, 
            text="NETEJAR MAPEIG DE NOMS", 
            command=self.reset_mapeig,
            bg="#ef4444",
            fg="white",
            active_bg="#dc2626",
            font_size=9,
            bold=False
        )
        self.btn_reset.pack(side="right", padx=(6, 0))
        
        self.progress = ttk.Progressbar(actions_card, orient="horizontal", mode="indeterminate")
        self.progress.pack(fill="x", pady=(0, 10))
        
        self.log_area = scrolledtext.ScrolledText(
            actions_card, 
            height=8,
            font=("Consolas", 9), 
            bg="#1e293b", 
            fg="#38bdf8", 
            insertbackground="white",
            relief="solid",
            bd=1
        )
        self.log_area.pack(fill="both", expand=True)

        # 4. Peu de pàgina (Footer) amb l'Autoria i Enllaç de Llicència
        footer_frame = tk.Frame(self.root, height=35, bg="#e2e8f0")
        footer_frame.pack(side="bottom", fill="x")
        
        autor = "Josepm"
        llicencia_text = "CC BY-NC-SA 4.0"
        url_llicencia = "https://creativecommons.org/licenses/by-nc-sa/4.0/"
        
        info_text = f"Autor: {autor}  |  Llicència: {llicencia_text}"
        
        footer_label = tk.Label(
            footer_frame,
            text=info_text,
            fg="#2563eb",
            cursor="hand2",
            bg="#e2e8f0",
            font=("Helvetica", 9, "bold"),
            wraplength=500
        )
        
        # Subratllar el text
        font_subratllada = font.Font(footer_label, footer_label.cget("font"))
        font_subratllada.configure(underline=True)
        footer_label.configure(font=font_subratllada)
        
        # Event del clic
        footer_label.bind("<Button-1>", lambda e: webbrowser.open_new(url_llicencia))
        footer_label.pack(pady=8)

    # --- Thread-Safe Queue helpers ---
    
    def check_queue(self):
        try:
            while True:
                func, args, kwargs = self.gui_queue.get_nowait()
                try:
                    func(*args, **kwargs)
                except Exception as e:
                    print(f"Error executant callback: {e}")
        except queue.Empty:
            pass
        self.root.after(50, self.check_queue)
        
    def run_in_gui(self, func, *args, **kwargs):
        self.gui_queue.put((func, args, kwargs))

    def log(self, missatge):
        self.run_in_gui(self._safe_log, missatge)
        
    def _safe_log(self, missatge):
        self.log_area.insert(tk.END, f"• {missatge}\n")
        self.log_area.see(tk.END)

    # --- Lògica d'arxius ---

    def seleccionar_fitxer(self):
        fitxer = filedialog.askopenfilename(
            title="Selecciona el fitxer Excel a processar",
            filetypes=[("Fitxers Excel", "*.xlsx *.xls"), ("Tots els fitxers", "*.*")]
        )
        if fitxer: 
            self.excel_file_path.set(fitxer)
            self.log(f"Fitxer Excel seleccionat: {fitxer}")

    def carregar_mapeig(self):
        if os.path.exists(FITXER_MAPPEIG):
            try:
                with open(FITXER_MAPPEIG, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                return {}
        return {}

    def guardar_mapeig(self):
        with open(FITXER_MAPPEIG, "w", encoding="utf-8") as f:
            json.dump(self.mapeig, f, ensure_ascii=False, indent=4)

    def reset_mapeig(self):
        if messagebox.askyesno("Confirmar", "Segur que vols esborrar el mapeig de noms?"):
            self.mapeig = {}
            if os.path.exists(FITXER_MAPPEIG): 
                os.remove(FITXER_MAPPEIG)
            self.log("♻️ Mapeig de noms reiniciat de forma permanent.")

    def obtenir_pseudonim(self, nom_real, prefix="Alumne"):
        nom_net = str(nom_real).strip()
        if not nom_net or nom_net == "-" or len(nom_net) < 3: 
            return nom_real
        
        # Normalització per evitar duplicats per diferències de majúscules/minúscules, accents o espais consecutius
        def normalitzar(s):
            s = str(s).lower().strip()
            s = re.sub(r'\s+', ' ', s)
            return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
            
        nom_norm = normalitzar(nom_net)
        
        # Cerquem si ja hi ha algun nom real al mapeig que equivalgui al nom normalitzat
        pseudonim = None
        for real_name, pseudo in self.mapeig.items():
            if normalitzar(real_name) == nom_norm:
                pseudonim = pseudo
                break
                
        if not pseudonim:
            # Comptar quants pseudònims amb aquest prefix hi ha registrats actualment
            count = sum(1 for pseudo in self.mapeig.values() if str(pseudo).startswith(prefix))
            num = count + 1
            pseudonim = f"{prefix} {num:03d}"
            self.mapeig[nom_net] = pseudonim
            self.guardar_mapeig()
            
        return pseudonim

    def obtenir_nom_real(self, pseudonim):
        pseudo_net = str(pseudonim).strip()
        for real, pseudo in self.mapeig.items():
            if pseudo == pseudo_net:
                return real
        return pseudonim

    def start_processing(self):
        # Validar ruta
        path = self.excel_file_path.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("Atenció", "Selecciona un fitxer Excel vàlid primer.")
            return
            
        # Validar columnes
        col_str = self.columna_noms.get().strip()
        if not col_str:
            messagebox.showwarning("Atenció", "Introdueix almenys una columna (ex: B).")
            return
            
        # Parsejar múltiples columnes separades per comes
        cols_raw = [c.strip().upper() for c in col_str.split(",")]
        col_indices = []
        for col in cols_raw:
            try:
                col_idx = column_index_from_string(col)
                col_indices.append(col_idx)
            except Exception:
                messagebox.showerror("Error", f"La columna '{col}' no és vàlida.")
                return
                
        # Parsejar prefixos de columnes separats per comes
        prefixes_str = self.prefixos_var.get().strip()
        prefixes = [p.strip() for p in prefixes_str.split(",")]
        
        # Si hi ha menys prefixos que columnes, omplim amb "Alumne" com a defecte segur
        while len(prefixes) < len(col_indices):
            prefixes.append("Alumne")
            
        # Paraula clau
        keyword = self.paraula_clau.get().lower().strip()
        is_reverse = self.mode_reves.get()
        
        if is_reverse and not self.mapeig:
            messagebox.showerror("Error", "No hi ha dades de mapeig carregades. No es pot desanonimitzar.")
            return
            
        # Modificar estats de la interfície
        self.btn_run.config(state="disabled")
        self.btn_browse.config(state="disabled")
        self.btn_reset.config(state="disabled")
        self.progress.start(12)
        self.log_area.delete("1.0", tk.END)
        self.log("S'està iniciant el processament multi-full...")

        # Fil secundari
        thread = threading.Thread(
            target=self.run_processing_worker,
            args=(path, col_indices, prefixes, keyword, is_reverse),
            daemon=True
        )
        thread.start()

    def run_processing_worker(self, path, col_indices, prefixes, keyword, is_reverse):
        import traceback
        try:
            dir_name = os.path.dirname(path)
            base_name = os.path.basename(path)
            
            if is_reverse:
                prefix_file = "REAL_"
                label_op = "Desanonimitzant"
            else:
                prefix_file = "ANON_"
                label_op = "Anonimitzant"
                
            ruta_sortida = os.path.join(dir_name, f"{prefix_file}{base_name}")
            
            self.log(f"Obrint el llibre Excel: {base_name}...")
            wb = openpyxl.load_workbook(path)
            
            sheet_names = wb.sheetnames
            self.log(f"S'han detectat {len(sheet_names)} fulls: {sheet_names}")
            
            canvis_totals = 0
            
            for sheet_name in sheet_names:
                self.log(f"Processant full: {sheet_name}...")
                ws = wb[sheet_name]
                
                # 1. Trobar la fila límit si s'especifica una paraula clau (només en mode anonimitzar)
                fila_limit = None
                if keyword:
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and keyword in str(cell.value).lower():
                                fila_limit = cell.row
                                break
                        if fila_limit: 
                            break
                
                if not is_reverse and fila_limit:
                    max_row = ws.max_row
                    num_a_esborrar = max_row - fila_limit
                    if num_a_esborrar > 0:
                        ws.delete_rows(fila_limit + 1, num_a_esborrar)
                        self.log(f"  🗑️ [{sheet_name}] Eliminades les files per sota de la paraula clau '{keyword}' (fila {fila_limit}).")

                # 2. Processar columnes indicades amb els seus prefixos associats
                limit = fila_limit if fila_limit else ws.max_row
                canvis_full = 0
                
                for r in range(1, limit + 1):
                    # Recórrer paral·lelament les columnes i els seus prefixos
                    for col_idx, prefix in zip(col_indices, prefixes):
                        cell = ws.cell(row=r, column=col_idx)
                        val = cell.value
                        if val:
                            # Si és revers, desanonimitzem
                            if is_reverse:
                                nou_val = self.obtenir_nom_real(val)
                            else:
                                # En anonimitzar, no anonimitzem la paraula clau en si
                                if keyword and keyword in str(val).lower():
                                    nou_val = val
                                else:
                                    nou_val = self.obtenir_pseudonim(val, prefix)
                                    
                            if nou_val != val:
                                cell.value = nou_val
                                canvis_full += 1
                                
                canvis_totals += canvis_full
                self.log(f"  ✅ [{sheet_name}] Full completat amb {canvis_full} canvis.")
            
            self.log(f"Desant el llibre Excel modificat a: {ruta_sortida}...")
            wb.save(ruta_sortida)
            self.log(f"🎉 Fet! Procés finalitzat amb èxit. {canvis_totals} canvis totals realitzats.")
            
            self.run_in_gui(self.on_processing_finished, True, f"Operació completada amb èxit a tots els fulls.\nFitxer desat a: {ruta_sortida}")
            
        except Exception as e:
            self.log(f"❌ Error en processar el fitxer: {e}")
            self.log(traceback.format_exc())
            self.run_in_gui(self.on_processing_finished, False, f"S'ha produït un error:\n{e}")

    def on_processing_finished(self, success, msg):
        self.progress.stop()
        self.btn_run.config(state="normal")
        self.btn_browse.config(state="normal")
        self.btn_reset.config(state="normal")
        if success:
            messagebox.showinfo("Èxit", msg)
        else:
            messagebox.showerror("Error", msg)


if __name__ == "__main__":
    root = tk.Tk()
    app = AnonimitzadorExcelApp(root)
    root.mainloop()
